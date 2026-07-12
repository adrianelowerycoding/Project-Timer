#### Written by Adrian Lowery ####

import os
import json
import datetime
import openpyxl
import tkinter as tk

from collections import defaultdict
from tkinter import ttk
from tkinter import filedialog, simpledialog, messagebox
from pathlib import Path

#### Creating the GUI ####

# -- Dark mode palette: tweak these hex values to taste --
BG = "#1e1e1e"          # main window / frame background
BG_ALT = "#2b2b2b"      # slightly lighter background (boxes, buttons)
FG = "#e0e0e0"          # primary text color
FG_MUTED = "#a0a0a0"    # secondary/less prominent text
ACCENT = "#4CAF50"      # accent color (e.g. idle circle button)
RUNNING = "#c0392b"     # accent color while a timer is running
BORDER = "#3c3c3c"      # borders/separators
DELETE_RED = "#c0392b"  # color for the "x" delete button


def format_hhmmss(total_seconds):
    """Format a number of seconds (int or float) as HH:MM:SS."""
    total_seconds = int(round(total_seconds))
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"


# The timer you create per project
class TimerBox(tk.Frame):
    def __init__(self, parent, project_name, project_number, on_delete, on_log_entry, **kwargs):
        # A slightly bigger frame so it reads as a "medium-sized square" button
        super().__init__(
            parent, bd=1, relief="solid", padx=10, pady=10,
            bg=BG_ALT, highlightbackground=BORDER,
            width=150, height=180, **kwargs
        )
        # Keep the frame from shrinking to fit its children so it stays square-ish
        self.pack_propagate(False)
        self.grid_propagate(False)

        self.project_name = project_name
        self.project_number = project_number
        self.on_delete = on_delete          # callback: on_delete(self)
        self.on_log_entry = on_log_entry    # callback: on_log_entry(entry_dict)

        # -- timer state --
        self.running = False
        self.start_time = None
        self._tick_job = None
        # locked while MainWindow is in "delete mode" so a click can't
        # accidentally start/stop a timer while the user is trying to delete one
        self.locked = False

        # -- "x" delete button, placed in the top-right corner --
        # Hidden by default; MainWindow toggles visibility via show_delete_x()/hide_delete_x()
        self.delete_btn = tk.Button(
            self, text="x", font=("Segoe UI", 8, "bold"),
            bg=DELETE_RED, fg="white", activebackground="#e74c3c", activeforeground="white",
            relief="flat", highlightthickness=0, bd=0,
            width=2, height=1, command=self.confirm_delete
        )
        # place() lets it float in the corner without disturbing the pack layout below

        self.title_label = tk.Label(
            self, text=project_name, font=("Segoe UI", 11, "bold"),
            bg=BG_ALT, fg=FG, wraplength=110, justify="center"
        )
        self.title_label.pack(pady=(0, 2))

        self.number_label = tk.Label(
            self, text=f"#{project_number}", font=("Segoe UI", 8),
            bg=BG_ALT, fg=FG_MUTED
        )
        self.number_label.pack(pady=(0, 4))

        self.canvas = tk.Canvas(self, width=70, height=70, highlightthickness=0, bg=BG_ALT)
        self.canvas.pack()

        self.circle = self.canvas.create_oval(4, 4, 66, 66, fill=ACCENT, outline="")
        self.circle_text = self.canvas.create_text(35, 35, text="Start", fill="white", font=("Segoe UI", 10, "bold"))
        self.canvas.bind("<Button-1>", self.on_toggle_clicked)

        self.time_label = tk.Label(self, text="00:00:00", font=("Consolas", 10), bg=BG_ALT, fg=FG_MUTED)
        self.time_label.pack(pady=(6, 0))

    # -- Start/stop logic --

    def on_toggle_clicked(self, event=None):
        if self.locked:
            return  # ignore clicks while delete mode is active
        if self.running:
            self.stop_timer()
        else:
            self.start_timer()

    def start_timer(self):
        self.running = True
        self.start_time = datetime.datetime.now()

        self.canvas.itemconfig(self.circle, fill=RUNNING)
        self.canvas.itemconfig(self.circle_text, text="Stop")

        self._tick()

    def _tick(self):
        # Called once a second while running, to keep the on-screen clock live
        if not self.running:
            return
        elapsed = datetime.datetime.now() - self.start_time
        self.time_label.config(text=self._format_duration(elapsed))
        self._tick_job = self.after(1000, self._tick)

    def stop_timer(self):
        if not self.running:
            return

        self.running = False
        if self._tick_job is not None:
            self.after_cancel(self._tick_job)
            self._tick_job = None

        end_time = datetime.datetime.now()
        duration = end_time - self.start_time

        # Reset visuals
        self.canvas.itemconfig(self.circle, fill=ACCENT)
        self.canvas.itemconfig(self.circle_text, text="Start")
        self.time_label.config(text="00:00:00")

        entry = {
            "project_name": self.project_name,
            "project_number": self.project_number,
            "start": self.start_time.strftime("%Y-%m-%d %H:%M:%S"),
            "end": end_time.strftime("%Y-%m-%d %H:%M:%S"),
            "duration_seconds": int(duration.total_seconds()),
        }
        self.start_time = None

        # Hand the finished entry straight up to MainWindow's in-memory session log
        self.on_log_entry(entry)

    @staticmethod
    def _format_duration(td):
        return format_hhmmss(td.total_seconds())

    # -- Delete mode / deletion --

    def show_delete_x(self):
        self.locked = True
        # Anchor to the top-right corner of the frame
        self.delete_btn.place(relx=1.0, x=-2, y=2, anchor="ne")

    def hide_delete_x(self):
        self.locked = False
        self.delete_btn.place_forget()

    def confirm_delete(self):
        if self.running:
            messagebox.showwarning(
                title="Timer Running",
                message=f'"{self.project_name}" is currently running. Stop it before deleting.'
            )
            return

        answer = messagebox.askyesno(
            title="Delete Timer",
            message=f'Do you really want to delete "{self.project_name}" (#{self.project_number})\'s timer?'
        )
        if answer:
            self.on_delete(self)


# The main window
class MainWindow(tk.Tk):

    # __init__ gets called as soon as mainWindow() object is called
    def __init__(self):
        super().__init__()
        self.title("Project Time Tracker")
        self.geometry("480x600")
        self.minsize(400, 500)
        self.configure(bg=BG)
        self.current_database = None       # display name, e.g. "john_Project_Tracker_db_123.xlsx"
        self.current_database_path = None  # full Path to the .xlsx file on disk
        self.current_database_name = tk.StringVar(value="Current Database: None")  # Initial text

        # In-memory data for the day. Every stop-click auto-fills a new entry here.
        self.session_log = []     # list of entry dicts, one per stop-click

        # project_number -> project_name, mirrored to a JSON sidecar file next to
        # the .xlsx so Load Database can rebuild timers with their real names
        self.project_names = {}

        # ttk widgets (just the Separators here) need a Style, since bg/fg
        # kwargs don't work on them directly like they do on plain tk widgets.
        style = ttk.Style(self)
        style.theme_use("clam")  # 'clam' respects custom colors; 'default'/'vista' mostly ignore them
        style.configure("TSeparator", background=BG)

        self.COLUMNS = 3

        # Keeps track of every TimerBox currently on screen
        self.timers = []
        # Whether the "x" delete buttons are currently showing on the timers
        self.delete_mode = False

        # Runs every single function inside mainWindow()
        # Creates the buttons and registers their references/callbacks
        self._build_section1()
        self._build_section2()
        self._build_title_bar()
        self._build_section3()

    # -- SECTION 1: Create Database / Load Database / View Day's Data / Update Database --
    # Creates the button(s) and hands Tkinter a reference to run logic once the button is pressed
    def _build_section1(self):
        frame = tk.Frame(self, pady=10, bg=BG)
        frame.pack(fill="x", padx=10)

        create_db_btn = tk.Button(
            frame, text="Create Database", command=self.on_create_database,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        create_db_btn.pack(side="left", padx=(0, 10))
        self.create_db_btn = create_db_btn  # kept so we can disable it during delete mode

        load_db_btn = tk.Button(
            frame, text="Load Database", command=self.on_load_database,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        load_db_btn.pack(side="left", padx=(0, 10))
        self.load_db_btn = load_db_btn  # kept so we can disable it during delete mode

        view_days_btn = tk.Button(
            frame, text="View Day's Data", command=self.on_view_days_data,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        view_days_btn.pack(side="left", padx=(0, 10))
        self.view_days_btn = view_days_btn  # kept so we can disable it during delete mode

        update_db_btn = tk.Button(
            frame, text="Update Database", command=self.on_update_database,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        update_db_btn.pack(side="left")
        self.update_db_btn = update_db_btn  # kept so we can disable it during delete mode

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=5, pady=(0, 5))

    # -- SECTION 2: Create/Delete timer buttons --
    def _build_section2(self):
        frame = tk.Frame(self, pady=5, bg=BG)
        frame.pack(fill="x", padx=10)

        create_timer_btn = tk.Button(
            frame, text="Create Project Timer", command=self.on_create_project_timer,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        create_timer_btn.pack(side="left", padx=(0, 10))
        self.create_timer_btn = create_timer_btn  # kept so we can disable it during delete mode

        # Keep a reference to this button so we can re-label it Done/Delete Timer
        self.delete_timer_btn = tk.Button(
            frame, text="Delete Timer", command=self.on_delete_timer,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        self.delete_timer_btn.pack(side="left")

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=5, pady=(5, 5))

    # -- Dynamic title bar: shows current database name --
    def _build_title_bar(self):
        label = tk.Label(
            self, textvariable=self.current_database_name,
            font=("Segoe UI", 9, "italic"), bg=BG, fg=FG_MUTED
        )
        label.pack(fill="x", padx=10, pady=(0, 5))

    # -- SECTION 3: Scrollable box of timer widgets --
    def _build_section3(self):
        outer = tk.Frame(self, bg=BG)
        outer.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.canvas = tk.Canvas(outer, highlightthickness=0, bg=BG)
        scrollbar = tk.Scrollbar(outer, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.timers_container = tk.Frame(self.canvas, bg=BG)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.timers_container, anchor="nw")

        self.timers_container.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))

    # -- Callback stubs --
    # All of the logic behind the buttons live in these callback functions

    # User creates database
    def on_create_database(self):

        # User inputs all project #s going into the database
        user_input_project_numbers = simpledialog.askstring(
            title="Input Project Numbers",
            prompt="Enter all project numbers included for this database (sep. by commas):")

        if user_input_project_numbers is None:
            return  # Cancelled - stop here, don't open save dialog
        if user_input_project_numbers.strip() == "":
            return  # Input was blank - re-prompt instead of just quitting

        # Strips the commas and spaces from project numbers
        user_input_project_numbers = user_input_project_numbers.replace(",", "_").replace(" ", "")

        # Captures username and sets path for suggested_folder and creates it
        username = os.environ.get("USERNAME", "user")  # Captures Window's username
        suggested_folder = Path(os.environ.get("USERPROFILE", r"C:\Users\Default")) / "Documents" / "Project Tracker" / f"db {user_input_project_numbers}"
        suggested_folder.mkdir(parents=True, exist_ok=True)  # creates folder

        # Opens File Explorer into created folder and offers save options
        filepath = filedialog.asksaveasfilename(
            title="Choose location for your Time Tracking database",
            initialdir=suggested_folder,
            initialfile=f"{username}_Project_Tracker_db_{user_input_project_numbers}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")]
        )

        if not filepath:
            return None

        filepath = Path(filepath)

        # Build the workbook itself: two static tabs in the same file, headers only,
        # no formulas. "Daily Log" holds one row per start/stop entry. "Project Totals"
        # holds one row per project per day. Both are appended to by "Update Database".
        wb = openpyxl.Workbook()
        ws_log = wb.active
        ws_log.title = "Daily Log"
        ws_log.append(["Project Name", "Project Number", "Start", "End", "Duration"])

        ws_totals = wb.create_sheet("Project Totals")
        ws_totals.append(["Project Name", "Project Number", "Total", "Date"])

        wb.save(filepath)

        # Setting Current Database Name and sending it to set_current_database() so it can
        # display current database name in the Title
        self.current_database = filepath.name
        self.current_database_path = filepath
        self.set_current_database(self.current_database)  # Passes database name into method

        # Fresh database means a fresh day's session. Any timers already on screen
        # get captured into the new sidecar right away.
        self.session_log = []
        self.project_names = {t.project_number: t.project_name for t in self.timers}
        self._save_sidecar()

        return filepath

    # Updates the "Curent Database: <database>" title from 'self.current_database' in
    # 'on_create_database()'. Showcases name in Title
    def set_current_database(self, name):
        self.current_database_name.set(f"Current Database: {name}")

    # -- JSON sidecar: same folder, same filename, ".json" instead of ".xlsx".
    #    Holds {project_number: project_name} so timer names survive a reload.
    #    This is just a plain text file next to the workbook - not a database
    #    engine, still nothing but openpyxl + stdlib. --
    def _sidecar_path(self):
        if not self.current_database_path:
            return None
        return Path(self.current_database_path).with_suffix(".json")

    def _save_sidecar(self):
        path = self._sidecar_path()
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.project_names, f, indent=2, sort_keys=True)
        except OSError:
            pass  # sidecar is a convenience file, not the source of truth for hours worked

    # User loads a previously created database to pick up where they left off.
    # Real names/numbers come from the JSON sidecar file saved next to the workbook.
    # If that sidecar is missing (an older database, or the file got separated),
    # we fall back to guessing project numbers out of the filename, same as before -
    # those timers just won't have a real name until re-typed.
    def on_load_database(self):
        base_folder = Path(os.environ.get("USERPROFILE", r"C:\Users\Default")) / "Documents" / "Project Tracker"
        base_folder.mkdir(parents=True, exist_ok=True)

        filepath = filedialog.askopenfilename(
            title="Select a Project Tracker database",
            initialdir=base_folder,
            filetypes=[("Excel file", "*.xlsx")]
        )

        if not filepath:
            return

        filepath = Path(filepath)

        # Sanity check: make sure this is actually one of our database files
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
        except Exception:
            messagebox.showerror(
                title="Load Database",
                message="That file couldn't be opened as an Excel database."
            )
            return

        if "Daily Log" not in sheet_names or "Project Totals" not in sheet_names:
            messagebox.showerror(
                title="Load Database",
                message="That file doesn't look like a Project Tracker database "
                        '(missing the "Daily Log" / "Project Totals" tabs).'
            )
            return

        # Warn before wiping out whatever's currently on screen
        if self.timers or self.session_log:
            proceed = messagebox.askyesno(
                title="Load Database",
                message="Loading a database replaces your current timers and any "
                        "unsaved time entries. Continue?"
            )
            if not proceed:
                return

        # Prefer the JSON sidecar - it has real names, not just numbers
        sidecar_path = filepath.with_suffix(".json")
        loaded_names = {}
        used_fallback = False

        if sidecar_path.exists():
            try:
                with open(sidecar_path, "r", encoding="utf-8") as f:
                    loaded_names = json.load(f)
            except (OSError, json.JSONDecodeError):
                loaded_names = {}

        if not loaded_names:
            # Fall back to parsing project numbers out of the filename, e.g.
            # "john_Project_Tracker_db_101_202_303.xlsx" -> ["101", "202", "303"]
            # The number becomes the name too, since that's all we have.
            stem = filepath.stem
            project_numbers_part = stem.split("_db_", 1)[1] if "_db_" in stem else ""
            loaded_names = {p: p for p in project_numbers_part.split("_") if p}
            used_fallback = True

        # Clear out current timers/session, exiting delete mode first if it's on
        if self.delete_mode:
            self.on_delete_timer()
        for timer in list(self.timers):
            timer.destroy()
        self.timers = []
        self.session_log = []

        # Recreate one timer per project number/name pair
        for number, name in loaded_names.items():
            timer = TimerBox(
                self.timers_container, name, number,
                on_delete=self.remove_timer,
                on_log_entry=self.add_log_entry
            )
            self.timers.append(timer)
        self._relayout_timers()

        self.current_database_path = filepath
        self.current_database = filepath.name
        self.set_current_database(self.current_database)
        self.project_names = dict(loaded_names)

        if not loaded_names:
            messagebox.showinfo(
                title="Load Database",
                message="Database loaded, but no project names or numbers could be "
                        "recovered. Use Create Project Timer to add timers manually."
            )
        elif used_fallback:
            messagebox.showinfo(
                title="Load Database",
                message="No name file (.json) was found next to this database, so "
                        "timer names were guessed from the filename and are just the "
                        "project numbers for now. A name file has been created going forward."
            )

        # Write (or (re)create) the sidecar now that we know it's in sync
        self._save_sidecar()

    # -- "Update Database": the only place Excel gets touched. Appends the day's
    #    in-memory session_log as static rows to the "Daily Log" tab, and today's
    #    per-project totals as static rows to the "Project Totals" tab - same file,
    #    two tabs. Nothing here recalculates or links back to the live session. --
    def on_update_database(self):
        if not self.current_database_path:
            messagebox.showwarning(
                title="No Database",
                message="Please create a database first (Create Database)."
            )
            return

        if not self.session_log:
            messagebox.showinfo(
                title="Update Database",
                message="No new time entries to save yet."
            )
            return

        today = datetime.date.today().isoformat()
        totals_seconds = defaultdict(int)
        totals_names = {}
        for entry in self.session_log:
            key = entry["project_number"]
            totals_seconds[key] += entry["duration_seconds"]
            totals_names[key] = entry["project_name"]

        wb = openpyxl.load_workbook(self.current_database_path)

        ws_log = wb["Daily Log"]
        for entry in self.session_log:
            ws_log.append([
                entry["project_name"], entry["project_number"], entry["start"], entry["end"],
                format_hhmmss(entry["duration_seconds"])
            ])

        ws_totals = wb["Project Totals"]
        for number, secs in totals_seconds.items():
            ws_totals.append([totals_names[number], number, format_hhmmss(secs), today])

        wb.save(self.current_database_path)

        messagebox.showinfo(
            title="Update Database",
            message=f"Saved {len(self.session_log)} entr{'y' if len(self.session_log) == 1 else 'ies'} to the database."
        )

        # Committed to the archive - clear the in-memory session so it isn't saved twice
        self.session_log = []

    # -- "View Day's Data": read-only look at today's entries, pulled straight from
    #    the in-memory session_log (auto-filled on every stop-click). Update Database
    #    plays no part in this - nothing here touches Excel. --
    def on_view_days_data(self):
        today = datetime.date.today().isoformat()
        todays_entries = [e for e in self.session_log if e["start"].startswith(today)]

        win = tk.Toplevel(self)
        win.title(f"Data for {today}")
        win.configure(bg=BG)
        win.geometry("560x560")

        style = ttk.Style(win)
        style.theme_use("clam")
        style.configure("Treeview", background=BG_ALT, fieldbackground=BG_ALT, foreground=FG)
        style.configure("Treeview.Heading", background=BG, foreground=FG)

        # -- Section: today's log entries --
        tk.Label(
            win, text="Today's Log", font=("Segoe UI", 11, "bold"), bg=BG, fg=FG
        ).pack(anchor="w", padx=10, pady=(10, 4))

        log_columns = ("project_name", "project_number", "start", "end", "duration")
        log_headers = ("Project Name", "Project Number", "Start", "End", "Duration")

        log_tree = ttk.Treeview(win, columns=log_columns, show="headings", height=8)
        for col, label in zip(log_columns, log_headers):
            log_tree.heading(col, text=label)
            log_tree.column(col, width=100, anchor="center")
        log_tree.pack(fill="x", padx=10)

        if todays_entries:
            for entry in todays_entries:
                log_tree.insert("", "end", values=(
                    entry["project_name"], entry["project_number"], entry["start"], entry["end"],
                    format_hhmmss(entry["duration_seconds"])
                ))
        else:
            tk.Label(
                win, text="No entries recorded for today yet.",
                bg=BG, fg=FG_MUTED
            ).pack(anchor="w", padx=10, pady=4)

        ttk.Separator(win, orient="horizontal").pack(fill="x", padx=10, pady=12)

        # -- Section: project totals --
        tk.Label(
            win, text="Project Totals", font=("Segoe UI", 11, "bold"), bg=BG, fg=FG
        ).pack(anchor="w", padx=10, pady=(0, 4))

        totals_seconds = defaultdict(int)
        totals_names = {}
        for entry in todays_entries:
            key = entry["project_number"]
            totals_seconds[key] += entry["duration_seconds"]
            totals_names[key] = entry["project_name"]

        totals_columns = ("project_name", "project_number", "total", "date")
        totals_headers = ("Project Name", "Project Number", "Total for Today", "Date")

        totals_tree = ttk.Treeview(win, columns=totals_columns, show="headings", height=6)
        for col, label in zip(totals_columns, totals_headers):
            totals_tree.heading(col, text=label)
            totals_tree.column(col, width=130, anchor="center")
        totals_tree.pack(fill="x", padx=10, pady=(0, 10))

        if totals_seconds:
            for number, secs in sorted(totals_seconds.items(), key=lambda kv: totals_names[kv[0]]):
                totals_tree.insert("", "end", values=(totals_names[number], number, format_hhmmss(secs), today))
        else:
            tk.Label(
                win, text="No totals to show yet.",
                bg=BG, fg=FG_MUTED
            ).pack(anchor="w", padx=10, pady=4)

    # -- Creates a new TimerBox. Input format is "Name, Project Number" - the
    #    project number must be at least 6 digits, since it's the identifier the
    #    JSON sidecar and Excel both key off of. --
    def on_create_project_timer(self):
        user_input = simpledialog.askstring(
            title="New Project Timer",
            prompt="Enter as: Name, Project Number\n"
                   "(project number must be at least 6 digits)\n"
                   "e.g. Kitchen Remodel, 123456"
        )

        if user_input is None:
            return  # Cancelled
        if user_input.strip() == "":
            return  # Blank input, ignore

        if "," not in user_input:
            messagebox.showerror(
                title="Invalid Format",
                message="Please enter using the format: Name, Project Number\n"
                        "e.g. Kitchen Remodel, 123456"
            )
            return

        name_part, _, number_part = user_input.partition(",")
        project_name = name_part.strip()
        project_number = number_part.strip().replace(" ", "")

        if not project_name:
            messagebox.showerror(
                title="Invalid Format",
                message="Please include a project name before the comma."
            )
            return

        if not project_number.isdigit() or len(project_number) < 6:
            messagebox.showerror(
                title="Invalid Project Number",
                message="Project number must be at least 6 digits (numbers only)."
            )
            return

        if any(t.project_number == project_number for t in self.timers):
            messagebox.showerror(
                title="Duplicate Project Number",
                message=f"A timer for project number {project_number} already exists."
            )
            return

        timer = TimerBox(
            self.timers_container, project_name, project_number,
            on_delete=self.remove_timer,
            on_log_entry=self.add_log_entry
        )
        self.timers.append(timer)
        self._relayout_timers()

        # Keep the sidecar's name mapping current
        self.project_names[project_number] = project_name
        self._save_sidecar()

    # -- Called by a TimerBox every time a stop-click produces a finished entry.
    #    This is the auto-fill: session_log updates the instant a timer is stopped. --
    def add_log_entry(self, entry):
        self.session_log.append(entry)

    # -- Toggles delete mode: shows/hides the "x" on every timer box, and disables
    #    every other button so "Done Deleting" is the only way out --
    def on_delete_timer(self):
        self.delete_mode = not self.delete_mode

        for timer in self.timers:
            if self.delete_mode:
                timer.show_delete_x()
            else:
                timer.hide_delete_x()

        # Stored as 'self' attributes so now reachable from 'on_delete_timer()'
        other_state = tk.DISABLED if self.delete_mode else tk.NORMAL
        self.create_db_btn.config(state=other_state)
        self.load_db_btn.config(state=other_state)
        self.view_days_btn.config(state=other_state)
        self.create_timer_btn.config(state=other_state)
        self.update_db_btn.config(state=other_state)

        self.delete_timer_btn.config(text="Done Deleting" if self.delete_mode else "Delete Timer")

    # -- Called by a TimerBox once the user confirms deletion --
    def remove_timer(self, timer):
        if timer in self.timers:
            self.timers.remove(timer)
        timer.destroy()
        self._relayout_timers()

    # -- Arranges all current TimerBoxes into a grid inside timers_container --
    def _relayout_timers(self):
        for widget in self.timers_container.winfo_children():
            widget.grid_forget()

        for index, timer in enumerate(self.timers):
            row = index // self.COLUMNS
            col = index % self.COLUMNS
            timer.grid(row=row, column=col, padx=8, pady=8)


if __name__ == "__main__":
    mainWindow = MainWindow()
    # Begins an event loop, listening for OS-level events and triggering callback functions
    mainWindow.mainloop()