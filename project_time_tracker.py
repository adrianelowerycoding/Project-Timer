import os
import re
import json
import datetime
import openpyxl
import tkinter as tk

from collections import defaultdict
from tkinter import ttk
from tkinter import filedialog, simpledialog, messagebox
from pathlib import Path
from zoneinfo import ZoneInfo

#### Creating the GUI ####

# -- Dark mode palette: tweak these hex values to taste --
BG = "#1e1e1e"          # main window / frame background
BG_ALT = "#2b2b2b"      # slightly lighter background (boxes, buttons)
FG = "#e0e0e0"          # primary text color
FG_MUTED = "#a0a0a0"    # secondary/less prominent text
ACCENT = "#4CAF50"      # accent color (e.g. idle circle button)
RUNNING = "#c0392b"     # accent color while a timer is running
BORDER = "#3c3c3c"      # borders/separators
DELETE_RED = "#c0392b"  # color for the "x" delete button

# -- Profile options: shift labels shown in the dropdown, and the timezone
#    labels mapped to their real IANA zone names. IANA names (not raw UTC
#    offsets) so DST is handled automatically for the US-based folks, while
#    Vietnam (which doesn't observe DST) stays correct year-round. --
SHIFT_LABELS = ["1st Shift", "2nd Shift", "3rd Shift"]
TIMEZONE_OPTIONS = {
    "Eastern - New Jersey (America/New_York)": "America/New_York",
    "Central - Arkansas (America/Chicago)": "America/Chicago",
    "Vietnam (Asia/Ho_Chi_Minh)": "Asia/Ho_Chi_Minh",
}

# -- Shift windows are anchored to a single reference timezone - Central Time,
#    Arkansas (headquarters) - since that's how the shift schedule is
#    actually defined company-wide, not something each site invents locally.
#    A worker elsewhere experiences a DIFFERENT local clock window for the
#    same shift - e.g. 3rd shift's 10pm-7am Central lands in the middle of
#    the afternoon/evening in Vietnam - so these hours get converted
#    per-person with real zoneinfo math (not a fixed hour offset), which
#    stays correct across DST changes even though Vietnam doesn't observe
#    DST at all. --
CENTRAL_TZ = ZoneInfo("America/Chicago")
SHIFT_WINDOWS_CENTRAL = {
    1: (datetime.time(6, 0), datetime.time(14, 0)),   # 1st shift: 6:00 AM - 2:00 PM Central
    2: (datetime.time(14, 0), datetime.time(23, 0)),  # 2nd shift: 2:00 PM - 11:00 PM Central
    3: (datetime.time(22, 0), datetime.time(7, 0)),   # 3rd shift: 10:00 PM - 7:00 AM Central
}


def _shift_local_window(worker_tz, shift_number, central_anchor_date):
    """Convert the Central-time-defined window for `shift_number` - the
    specific occurrence anchored to `central_anchor_date` (a Central calendar
    date) - into the equivalent timezone-aware start/end datetimes in
    `worker_tz`. Uses real zoneinfo conversion rather than arithmetic on raw
    hours, so DST (and Vietnam's lack of it) comes out correct automatically."""
    start_t, end_t = SHIFT_WINDOWS_CENTRAL.get(shift_number, SHIFT_WINDOWS_CENTRAL[1])
    start_central = datetime.datetime.combine(central_anchor_date, start_t, tzinfo=CENTRAL_TZ)
    end_date = central_anchor_date if end_t > start_t else central_anchor_date + datetime.timedelta(days=1)
    end_central = datetime.datetime.combine(end_date, end_t, tzinfo=CENTRAL_TZ)
    return start_central.astimezone(worker_tz), end_central.astimezone(worker_tz)


def compute_shift_date(local_dt, shift_number):
    """Return the date (a datetime.date, in the WORKER's own local calendar)
    that an entire shift should be attributed to, given a timezone-aware
    local datetime and a shift number. This is the "Shift Date" used to
    group Project Totals.

    Shift hours are defined once, in Central time (SHIFT_WINDOWS_CENTRAL).
    Step 1 figures out which Central calendar day's occurrence of the shift
    this instant belongs to, using the "before the end time -> previous day"
    rule evaluated in Central time, since that's where the boundaries are
    actually defined. Step 2 converts THAT specific occurrence's start time
    into the worker's own local calendar - that converted date is what gets
    recorded. This stays correct even when a shift that doesn't cross
    midnight in Central time ends up crossing midnight once translated into
    a worker's own local time (or vice versa) - no special-casing needed,
    since the "crossing" behavior falls out naturally from converting one
    real, absolute instant into another timezone.
    """
    worker_tz = local_dt.tzinfo
    central_dt = local_dt.astimezone(CENTRAL_TZ)

    start_t, end_t = SHIFT_WINDOWS_CENTRAL.get(shift_number, SHIFT_WINDOWS_CENTRAL[1])
    crosses_midnight_central = end_t <= start_t
    if crosses_midnight_central and central_dt.time() < end_t:
        central_anchor_date = (central_dt - datetime.timedelta(days=1)).date()
    else:
        central_anchor_date = central_dt.date()

    start_local, _end_local = _shift_local_window(worker_tz, shift_number, central_anchor_date)
    return start_local.date()


def format_hhmmss(total_seconds):
    """Format a number of seconds (int or float) as HH:MM:SS. Negative values
    (e.g. an Overhead Total that's gone over budget) are shown as -HH:MM:SS."""
    total_seconds = int(round(total_seconds))
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{sign}{hours:02}:{minutes:02}:{seconds:02}"


def parse_hhmmss(text):
    """Reverse of format_hhmmss: 'HH:MM:SS' (optionally '-HH:MM:SS') -> seconds."""
    text = str(text).strip()
    sign = -1 if text.startswith("-") else 1
    text = text.lstrip("-")
    parts = text.split(":")
    if len(parts) != 3:
        return 0
    try:
        hours, minutes, seconds = (int(p) for p in parts)
    except ValueError:
        return 0
    return sign * (hours * 3600 + minutes * 60 + seconds)


# The timer you create per project
class TimerBox(tk.Frame):
    def __init__(self, parent, project_name, project_number, on_delete, on_log_entry, on_state_change, **kwargs):
        # A slightly bigger frame so it reads as a "medium-sized square" button
        super().__init__(
            parent, bd=1, relief="solid", padx=10, pady=10,
            bg=BG_ALT, highlightbackground=BORDER,
            width=150, height=180, **kwargs
        )
        # Keep the frame from shrinking to fit its children so it stays square-ish
        self.pack_propagate(False)
        self.grid_propagate(False)

        self.project_name = project_name
        self.project_number = project_number
        self.on_delete = on_delete            # callback: on_delete(self)
        self.on_log_entry = on_log_entry      # callback: on_log_entry(entry_dict)
        self.on_state_change = on_state_change  # callback: on_state_change(self, running)

        # -- timer state --
        self.running = False
        self.start_time = None
        self._tick_job = None
        # locked while MainWindow is in "delete mode" so a click can't
        # accidentally start/stop a timer while the user is trying to delete one
        self.locked = False

        # -- "x" delete button, placed in the top-right corner --
        # Hidden by default; MainWindow toggles visibility via show_delete_x()/hide_delete_x()
        self.delete_btn = tk.Button(
            self, text="x", font=("Segoe UI", 8, "bold"),
            bg=DELETE_RED, fg="white", activebackground="#e74c3c", activeforeground="white",
            relief="flat", highlightthickness=0, bd=0,
            width=2, height=1, command=self.confirm_delete
        )
        # place() lets it float in the corner without disturbing the pack layout below

        self.title_label = tk.Label(
            self, text=project_name, font=("Segoe UI", 11, "bold"),
            bg=BG_ALT, fg=FG, wraplength=110, justify="center"
        )
        self.title_label.pack(pady=(0, 2))

        self.number_label = tk.Label(
            self, text=f"#{project_number}", font=("Segoe UI", 8),
            bg=BG_ALT, fg=FG_MUTED
        )
        self.number_label.pack(pady=(0, 4))

        self.canvas = tk.Canvas(self, width=70, height=70, highlightthickness=0, bg=BG_ALT)
        self.canvas.pack()

        self.circle = self.canvas.create_oval(4, 4, 66, 66, fill=ACCENT, outline="")
        self.circle_text = self.canvas.create_text(35, 35, text="Start", fill="white", font=("Segoe UI", 10, "bold"))
        self.canvas.bind("<Button-1>", self.on_toggle_clicked)

        self.time_label = tk.Label(self, text="00:00:00", font=("Consolas", 10), bg=BG_ALT, fg=FG_MUTED)
        self.time_label.pack(pady=(6, 0))

    # -- Start/stop logic --

    def on_toggle_clicked(self, event=None):
        if self.locked:
            return  # ignore clicks while delete mode is active
        if self.running:
            self.stop_timer()
        else:
            self.start_timer()

    def start_timer(self):
        self.running = True
        # Captured in UTC (timezone-aware) rather than naive local time, so the
        # entry doesn't depend on the machine's OS timezone being set correctly -
        # MainWindow converts it to the user's profile timezone at log time.
        self.start_time = datetime.datetime.now(datetime.timezone.utc)

        self.canvas.itemconfig(self.circle, fill=RUNNING)
        self.canvas.itemconfig(self.circle_text, text="Stop")

        self._tick()
        self.on_state_change(self, True)

    def _tick(self):
        # Called once a second while running, to keep the on-screen clock live
        if not self.running:
            return
        elapsed = datetime.datetime.now(datetime.timezone.utc) - self.start_time
        self.time_label.config(text=self._format_duration(elapsed))
        self._tick_job = self.after(1000, self._tick)

    def stop_timer(self):
        if not self.running:
            return

        self.running = False
        if self._tick_job is not None:
            self.after_cancel(self._tick_job)
            self._tick_job = None

        end_time = datetime.datetime.now(datetime.timezone.utc)
        duration = end_time - self.start_time

        # Reset visuals
        self.canvas.itemconfig(self.circle, fill=ACCENT)
        self.canvas.itemconfig(self.circle_text, text="Start")
        self.time_label.config(text="00:00:00")

        # Raw UTC timestamps only - MainWindow.add_log_entry() resolves these
        # against the CURRENT profile (shift + timezone) and bakes in Shift,
        # TimeZone, and ShiftDate once, at the moment the entry is created.
        entry = {
            "project_name": self.project_name,
            "project_number": self.project_number,
            "start_utc": self.start_time,
            "end_utc": end_time,
            "duration_seconds": int(duration.total_seconds()),
        }
        self.start_time = None

        # Hand the finished entry straight up to MainWindow's in-memory session log
        self.on_log_entry(entry)
        self.on_state_change(self, False)

    @staticmethod
    def _format_duration(td):
        return format_hhmmss(td.total_seconds())

    # -- Delete mode / deletion --

    def show_delete_x(self):
        self.locked = True
        # Anchor to the top-right corner of the frame
        self.delete_btn.place(relx=1.0, x=-2, y=2, anchor="ne")

    def hide_delete_x(self):
        self.locked = False
        self.delete_btn.place_forget()

    def confirm_delete(self):
        if self.running:
            messagebox.showwarning(
                title="Timer Running",
                message=f'"{self.project_name}" is currently running. Stop it before deleting.'
            )
            return

        answer = messagebox.askyesno(
            title="Delete Timer",
            message=f'Do you really want to delete "{self.project_name}" (#{self.project_number})\'s timer?'
        )
        if answer:
            self.on_delete(self)


# The main window
class MainWindow(tk.Tk):

    # __init__ gets called as soon as mainWindow() object is called
    def __init__(self):
        super().__init__()
        self.title("Project Time Tracker")
        self.geometry("800x500")
        self.minsize(550, 325) #x, y
        self.configure(bg=BG)
        self.current_database = None       # display name, e.g. "john_Project_Tracker_db_123.xlsx"
        self.current_database_path = None  # full Path to the .xlsx file on disk
        self.current_database_name = tk.StringVar(value="Current Database: None")  # Initial text

        # In-memory data for the day. Every stop-click auto-fills a new entry here.
        self.session_log = []     # list of entry dicts, one per stop-click

        # project_number -> project_name, mirrored to a JSON sidecar file next to
        # the .xlsx so Load Database can rebuild timers with their real names
        self.project_names = {}

        # Profile settings (Shift + Timezone) for whoever is running this instance.
        # Lives in its own JSON file, independent of whichever database is loaded -
        # a person's shift/timezone doesn't change based on which project they're
        # tracking. Loaded once here; edited via Options > Profile.
        self.profile = self._load_profile()

        # ttk widgets (just the Separators here) need a Style, since bg/fg
        # kwargs don't work on them directly like they do on plain tk widgets.
        style = ttk.Style(self)
        style.theme_use("clam")  # 'clam' respects custom colors; 'default'/'vista' mostly ignore them
        style.configure("TSeparator", background=BG)

        # A ttk Combobox's dropdown list (the "popdown") is an internal plain
        # Tk Listbox that ttk.Style doesn't reach - it needs the classic Tk
        # option database instead. Without this, the highlighted/selected
        # item renders in light text on a light highlight and is basically
        # unreadable against the dark theme. Applied once here, at the root,
        # so every Combobox in the app (including ones in Toplevels, like
        # Profile) picks it up.
        self.option_add("*TCombobox*Listbox.background", BG_ALT)
        self.option_add("*TCombobox*Listbox.foreground", FG)
        self.option_add("*TCombobox*Listbox.selectBackground", ACCENT)
        self.option_add("*TCombobox*Listbox.selectForeground", BG)

        self.COLUMNS = 3

        # Keeps track of every TimerBox currently on screen
        self.timers = []
        # Whether the "x" delete buttons are currently showing on the timers
        self.delete_mode = False

        # Runs every single function inside mainWindow()
        # Creates the buttons and registers their references/callbacks
        self._build_section1()
        self._build_section2()
        self._build_title_bar()
        self._build_section3()

        # No database loaded yet - Create/Delete Timer start greyed out
        self._refresh_timer_controls()

        # Catch the window's own close button ("X") so we can warn about unsaved work
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    # -- SECTION 1: Create Database / Load Database / View Day's Data / Update Database / Options --
    # Creates the button(s) and hands Tkinter a reference to run logic once the button is pressed
    def _build_section1(self):
        frame = tk.Frame(self, pady=10, bg=BG)
        frame.pack(fill="x", padx=10)

        create_db_btn = tk.Button(
            frame, text="Create Database", command=self.on_create_database,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        create_db_btn.pack(side="left", padx=(0, 10))
        self.create_db_btn = create_db_btn  # kept so we can disable it during delete mode

        load_db_btn = tk.Button(
            frame, text="Load Database", command=self.on_load_database,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        load_db_btn.pack(side="left", padx=(0, 10))
        self.load_db_btn = load_db_btn  # kept so we can disable it during delete mode

        view_days_btn = tk.Button(
            frame, text="View Day's Data", command=self.on_view_days_data,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        view_days_btn.pack(side="left", padx=(0, 10))
        self.view_days_btn = view_days_btn  # kept so we can disable it during delete mode

        update_db_btn = tk.Button(
            frame, text="Update Database", command=self.on_update_database,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        update_db_btn.pack(side="left")
        self.update_db_btn = update_db_btn  # kept so we can disable it during delete mode

        # -- Options dropdown: currently just "Profile", room to grow later --
        options_mb = tk.Menubutton(
            frame, text="Options", bg=BG_ALT, fg=FG,
            activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        options_menu = tk.Menu(
            options_mb, tearoff=0, bg=BG_ALT, fg=FG,
            activebackground=BORDER, activeforeground=FG
        )
        options_menu.add_command(label="Profile", command=self.on_open_profile)
        options_mb.config(menu=options_menu)
        options_mb.pack(side="left", padx=(10, 0))
        self.options_mb = options_mb  # kept so we can disable it during delete mode / running timers

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=5, pady=(0, 5))

    # -- SECTION 2: Create/Delete timer buttons --
    def _build_section2(self):
        frame = tk.Frame(self, pady=5, bg=BG)
        frame.pack(fill="x", padx=10)

        create_timer_btn = tk.Button(
            frame, text="Create Project Timer", command=self.on_create_project_timer,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        create_timer_btn.pack(side="left", padx=(0, 10))
        self.create_timer_btn = create_timer_btn  # kept so we can disable it during delete mode

        # Keep a reference to this button so we can re-label it Done/Delete Timer
        self.delete_timer_btn = tk.Button(
            frame, text="Delete Timer", command=self.on_delete_timer,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0
        )
        self.delete_timer_btn.pack(side="left")

        ttk.Separator(self, orient="horizontal").pack(fill="x", padx=5, pady=(5, 5))

    # -- Dynamic title bar: shows current database name --
    def _build_title_bar(self):
        label = tk.Label(
            self, textvariable=self.current_database_name,
            font=("Segoe UI", 9, "italic"), bg=BG, fg=FG_MUTED
        )
        label.pack(fill="x", padx=10, pady=(0, 5))

    # -- SECTION 3: Scrollable box of timer widgets --
    def _build_section3(self):
        outer = tk.Frame(self, bg=BG)
        outer.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.canvas = tk.Canvas(outer, highlightthickness=0, bg=BG)
        scrollbar = tk.Scrollbar(outer, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        self.timers_container = tk.Frame(self.canvas, bg=BG)
        self.canvas_window = self.canvas.create_window((0, 0), window=self.timers_container, anchor="nw")

        self.timers_container.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Configure>", lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width))

    # -- Callback stubs --
    # All of the logic behind the buttons live in these callback functions

    # -- Unsaved-work guard, shared by the window close button, Create Database,
    #    and Load Database. Returns True if it's OK to proceed (nothing unsaved,
    #    or the user confirmed they're fine discarding it). --
    def _confirm_discard_if_unsaved(self):
        if not self.session_log:
            return True
        if self.current_database:
            message = f"Are you sure you want to exit without saving {self.current_database}?"
        else:
            message = "Are you sure you want to exit without saving your current time entries?"
        return messagebox.askyesno(title="Unsaved Changes", message=message)

    # -- Called when the window's own close button ("X") is clicked --
    def on_close(self):
        if self._confirm_discard_if_unsaved():
            self.destroy()

    # -- Wipes the current timers/session/name-mapping so a new or freshly loaded
    #    database starts from a clean slate. --
    def _clear_workspace(self):
        if self.delete_mode:
            self.on_delete_timer()  # exit delete mode cleanly first
        for timer in list(self.timers):
            timer.destroy()
        self.timers = []
        self.session_log = []
        self.project_names = {}

    # -- Create Project Timer / Delete Timer only make sense once a database is
    #    loaded (that's what the sidecar writes get attached to). Grey them out
    #    otherwise. --
    def _refresh_timer_controls(self):
        state = tk.NORMAL if self.current_database_path else tk.DISABLED
        self.create_timer_btn.config(state=state)
        self.delete_timer_btn.config(state=state)

    # -- Profile: shift + timezone for whoever is running this instance --

    def _profile_path(self):
        folder = Path(os.environ.get("USERPROFILE", r"C:\Users\Default")) / "Documents" / "Project Tracker"
        folder.mkdir(parents=True, exist_ok=True)
        return folder / "profile.json"

    def _load_profile(self):
        path = self._profile_path()
        if path.exists():
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict) and "shift" in data and "timezone" in data:
                    return data
            except (OSError, json.JSONDecodeError):
                pass
        # Sane defaults until the user sets their own via Options > Profile
        return {"shift": 1, "timezone": "America/New_York"}

    def _save_profile(self):
        path = self._profile_path()
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.profile, f, indent=2, sort_keys=True)
        except OSError:
            pass  # profile is a convenience file; app still runs fine without it persisting

    def on_open_profile(self):
        win = tk.Toplevel(self)
        win.title("Profile")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()  # small modal-style window

        pstyle = ttk.Style(win)
        pstyle.theme_use("clam")
        pstyle.configure("TCombobox", fieldbackground=BG_ALT, background=BG_ALT, foreground=FG_MUTED)
        # The 'clam' theme maps its own fieldbackground/foreground for the
        # "readonly" state (which is what our Comboboxes use), overriding the
        # configure() call above once a value is selected and focus leaves the
        # box. Explicitly map the readonly state too, or the closed box falls
        # back to a light field with light text.
        pstyle.map(
            "TCombobox",
            fieldbackground=[("readonly", BG_ALT)],
            foreground=[("readonly", FG_MUTED)],
            selectforeground=[("readonly", FG_MUTED)],
            selectbackground=[("readonly", BG_ALT)],
        )

        current_shift_label = SHIFT_LABELS[self.profile.get("shift", 1) - 1]
        tz_labels = list(TIMEZONE_OPTIONS.keys())
        current_tz_value = self.profile.get("timezone", "America/New_York")
        current_tz_label = next(
            (label for label, value in TIMEZONE_OPTIONS.items() if value == current_tz_value),
            tz_labels[0]
        )

        # -- Status titles: show exactly what's currently selected, updated
        #    live as the dropdowns change, so it's never ambiguous what's
        #    about to be saved. --
        shift_status_var = tk.StringVar(value=f"Shift: {current_shift_label}")
        tz_status_var = tk.StringVar(value=f"Timezone: {current_tz_label}")

        tk.Label(
            win, textvariable=shift_status_var, font=("Segoe UI", 10, "bold"), bg=BG, fg=FG
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=12, pady=(14, 2))

        shift_var = tk.StringVar(value=current_shift_label)
        shift_box = ttk.Combobox(
            win, textvariable=shift_var, values=SHIFT_LABELS,
            state="readonly", width=24
        )
        shift_box.grid(row=1, column=0, columnspan=2, sticky="w", padx=12, pady=(0, 10))
        shift_box.bind("<<ComboboxSelected>>", lambda e: shift_status_var.set(f"Shift: {shift_var.get()}"))

        tk.Label(
            win, textvariable=tz_status_var, font=("Segoe UI", 10, "bold"), bg=BG, fg=FG
        ).grid(row=2, column=0, columnspan=2, sticky="w", padx=12, pady=(4, 2))

        tz_var = tk.StringVar(value=current_tz_label)
        tz_box = ttk.Combobox(
            win, textvariable=tz_var, values=tz_labels,
            state="readonly", width=34
        )
        tz_box.grid(row=3, column=0, columnspan=2, sticky="w", padx=12, pady=(0, 10))
        tz_box.bind("<<ComboboxSelected>>", lambda e: tz_status_var.set(f"Timezone: {tz_var.get()}"))

        btn_frame = tk.Frame(win, bg=BG)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=(10, 14))

        def save_and_close():
            self.profile["shift"] = SHIFT_LABELS.index(shift_var.get()) + 1
            self.profile["timezone"] = TIMEZONE_OPTIONS[tz_var.get()]
            self._save_profile()
            win.destroy()

        tk.Button(
            btn_frame, text="Save", command=save_and_close,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0, width=10
        ).pack(side="left", padx=6)

        tk.Button(
            btn_frame, text="Cancel", command=win.destroy,
            bg=BG_ALT, fg=FG, activebackground=BORDER, activeforeground=FG,
            relief="flat", highlightthickness=0, width=10
        ).pack(side="left", padx=6)

        win.update_idletasks()
        win.geometry(f"+{self.winfo_rootx() + 80}+{self.winfo_rooty() + 80}")

    # User creates database
    def on_create_database(self):

        # Switching databases while one's already loaded - make sure nothing
        # unsaved gets silently thrown away first
        if self.current_database_path and not self._confirm_discard_if_unsaved():
            return

        # User names this database after the project(s) it covers. Individual
        # project numbers are still attached per-timer via Create Project Timer -
        # this input just builds the file/folder name.
        user_input_project_names = simpledialog.askstring(
            title="Input Project Names",
            prompt="Input name of your database\nSuggested format 'Project1_Project2_Project3`")

        if user_input_project_names is None:
            return  # Cancelled - stop here, don't open save dialog
        if user_input_project_names.strip() == "":
            return  # Input was blank - re-prompt instead of just quitting

        # Commas become underscores, spaces are dropped entirely
        db_name = user_input_project_names.replace(",", "_").replace(" ", "")
        db_name = self._sanitize_for_filename(db_name) #Static method that strips chars. Windows doesn't allow
        if not db_name:
            messagebox.showerror(
                title="Invalid Name",
                message="Please enter a database name"
            )
            return

        # Captures username and sets path for suggested_folder and creates it
        username = os.environ.get("USERNAME", "user")  # Captures Window's username
        suggested_folder = Path(os.environ.get("USERPROFILE", r"C:\Users\Default")) / "Documents" / "Project Tracker" / f"db {db_name}"
        suggested_folder.mkdir(parents=True, exist_ok=True)  # creates folder

        # Opens File Explorer into created folder and offers save options
        filepath = filedialog.asksaveasfilename(
            title="Choose location for your Time Tracking database",
            initialdir=suggested_folder,
            initialfile=f"{username}_Project_Tracker_db_{db_name}.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")]
        )

        if not filepath:
            return None

        filepath = Path(filepath)

        # Build the workbook itself: two static tabs in the same file, headers only,
        # no formulas. "Daily Log" holds one row per start/stop entry. "Project Totals"
        # holds one row per project per day. Both are appended to by "Update Database".
        # The file is named for the project(s) it covers, but every row still carries
        # a Project Number column, so the data stays searchable/keyed by project number.
        wb = openpyxl.Workbook()
        ws_log = wb.active
        ws_log.title = "Daily Log"
        ws_log.append(["Project Name", "Project Number", "Start", "End", "Duration", "Shift", "Time Zone"])

        ws_totals = wb.create_sheet("Project Totals")
        ws_totals.append(["Project Name", "Project Number", "Total", "Shift Date", "Shift", "Time Zone"])

        wb.save(filepath)

        # New database means a clean slate - old timers/session don't carry over
        self._clear_workspace()

        # Setting Current Database Name and sending it to set_current_database() so it can
        # display current database name in the Title
        self.current_database = filepath.name
        self.current_database_path = filepath
        self.set_current_database(self.current_database)  # Passes database name into method
        self._save_sidecar()
        self._refresh_timer_controls()

        return filepath

    @staticmethod
    def _sanitize_for_filename(text):
        # Strip characters Windows doesn't allow in filenames
        return re.sub(r'[<>:"/\\|?*]', "", text)

    # Updates the "Curent Database: <database>" title from 'self.current_database' in
    # 'on_create_database()'. Showcases name in Title
    def set_current_database(self, name):
        self.current_database_name.set(f"Current Database: {name}")

    # -- JSON sidecar: same folder, same filename, ".json" instead of ".xlsx".
    #    Holds {project_number: project_name} so timer names survive a reload.
    #    This is just a plain text file next to the workbook - not a database
    #    engine, still nothing but openpyxl + stdlib. --
    def _sidecar_path(self):
        if not self.current_database_path:
            return None
        return Path(self.current_database_path).with_suffix(".json")

    def _save_sidecar(self):
        path = self._sidecar_path()
        if not path:
            return
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(self.project_names, f, indent=2, sort_keys=True)
        except OSError:
            pass  # sidecar is a convenience file, not the source of truth for hours worked

    @staticmethod
    def _ensure_headers(ws, expected_headers):
        """Backfill any missing header cells in row 1 against `expected_headers`.
        Databases created before the Shift/Time Zone columns existed already
        have data written into those columns (since column position is fixed
        by list order), just no header label - this patches the label in
        without touching any existing data."""
        for col_idx, header in enumerate(expected_headers, start=1):
            if not ws.cell(row=1, column=col_idx).value:
                ws.cell(row=1, column=col_idx, value=header)

    # User loads a previously created database to pick up where they left off.
    # Real names/numbers come from the JSON sidecar file saved next to the workbook -
    # the filename itself is just a human-readable label now, not an encoding of
    # project numbers, so there's no filename to fall back on parsing anymore.
    def on_load_database(self):
        # Switching databases while one's already loaded - make sure nothing
        # unsaved gets silently thrown away first
        if self.current_database_path and not self._confirm_discard_if_unsaved():
            return

        base_folder = Path(os.environ.get("USERPROFILE", r"C:\Users\Default")) / "Documents" / "Project Tracker"
        base_folder.mkdir(parents=True, exist_ok=True)

        filepath = filedialog.askopenfilename(
            title="Select a Project Tracker database",
            initialdir=base_folder,
            filetypes=[("Excel file", "*.xlsx")]
        )

        if not filepath:
            return

        filepath = Path(filepath)

        # Open once, writable - we validate the tabs AND combine any duplicate
        # (date, project number) rows in "Project Totals" left over from prior sessions.
        try:
            wb = openpyxl.load_workbook(filepath)
        except Exception:
            messagebox.showerror(
                title="Load Database",
                message="That file couldn't be opened as an Excel database."
            )
            return

        if "Daily Log" not in wb.sheetnames or "Project Totals" not in wb.sheetnames:
            wb.close()
            messagebox.showerror(
                title="Load Database",
                message="That file doesn't look like a Project Tracker database "
                        '(missing the "Daily Log" / "Project Totals" tabs).'
            )
            return

        # Heal headers on older databases created before the Shift/Time Zone
        # columns existed - the columns already had data once this program
        # started writing to them, just no label in row 1.
        self._ensure_headers(wb["Daily Log"], ["Project Name", "Project Number", "Start", "End", "Duration", "Shift", "Time Zone"])
        self._ensure_headers(wb["Project Totals"], ["Project Name", "Project Number", "Total", "Shift Date", "Shift", "Time Zone"])

        # -- Combine duplicate (shift date, project number) rows into one summed row --
        ws_totals = wb["Project Totals"]
        combined = {}  # (shift_date, project_number) -> {"name": ..., "seconds": ..., "shift": ..., "timezone": ...}
        for row in ws_totals.iter_rows(min_row=2, values_only=True):
            if not row or row[1] is None:
                continue
            # Pad to 6 columns so older files (saved before the Shift/Time
            # Zone columns existed) still load without crashing.
            project_name, project_number, total_str, shift_date_str, shift_value, timezone_value = (list(row) + [None] * 6)[:6]
            key = (shift_date_str, project_number)
            seconds = parse_hhmmss(total_str)
            if key in combined:
                combined[key]["seconds"] += seconds
                if not combined[key]["name"] and project_name:
                    combined[key]["name"] = project_name
                if not combined[key]["shift"] and shift_value:
                    combined[key]["shift"] = shift_value
                if not combined[key]["timezone"] and timezone_value:
                    combined[key]["timezone"] = timezone_value
            else:
                combined[key] = {"name": project_name, "seconds": seconds, "shift": shift_value, "timezone": timezone_value}

        if ws_totals.max_row > 1:
            ws_totals.delete_rows(2, ws_totals.max_row - 1)
        for (shift_date_str, project_number), data in sorted(combined.items(), key=lambda kv: (kv[0][0] or "", kv[0][1] or "")):
            ws_totals.append([
                data["name"], project_number, format_hhmmss(data["seconds"]), shift_date_str,
                data["shift"], data["timezone"]
            ])

        wb.save(filepath)
        wb.close()

        # Prefer the JSON sidecar - it has the real names, keyed by project number
        sidecar_path = filepath.with_suffix(".json")
        loaded_names = {}

        if sidecar_path.exists():
            try:
                with open(sidecar_path, "r", encoding="utf-8") as f:
                    loaded_names = json.load(f)
            except (OSError, json.JSONDecodeError):
                loaded_names = {}

        # Clear out the current workspace - old database's timers/session don't carry over
        self._clear_workspace()

        # Recreate one timer per project number/name pair
        for number, name in loaded_names.items():
            timer = TimerBox(
                self.timers_container, name, number,
                on_delete=self.remove_timer,
                on_log_entry=self.add_log_entry,
                on_state_change=self.handle_timer_state_change
            )
            self.timers.append(timer)
        self._relayout_timers()

        self.current_database_path = filepath
        self.current_database = filepath.name
        self.set_current_database(self.current_database)
        self.project_names = dict(loaded_names)
        self._refresh_timer_controls()

        if not loaded_names:
            messagebox.showinfo(
                title="Load Database",
                message="Database loaded, but no name file (.json) was found next to "
                        "it, so no timers could be rebuilt. Use Create Project Timer "
                        "to add them manually - a name file will be created from there."
            )

    # -- "Update Database": the only place Excel gets touched. Appends the day's
    #    in-memory session_log as static rows to the "Daily Log" tab. For
    #    "Project Totals", a row for the same (date, project number) gets its
    #    total added to rather than duplicated - so hitting Update Database more
    #    than once for the same project on the same day still yields one row. --
    def on_update_database(self):
        if not self.current_database_path:
            messagebox.showwarning(
                title="No Database",
                message="Please create a database first (Create Database)."
            )
            return

        if not self.session_log:
            messagebox.showinfo(
                title="Update Database",
                message="No new time entries to save yet."
            )
            return

        # Grouping key is (shift_date, project_number) - NOT the calendar date -
        # so a 3rd-shift session that starts at 10pm and runs past midnight
        # still lands in a single Project Totals row, dated to the day the
        # shift started on. Each entry already carries its own shift_date/shift,
        # baked in by add_log_entry() when it was created.
        totals_seconds = defaultdict(int)
        totals_names = {}
        totals_shift = {}
        totals_timezone = {}
        for entry in self.session_log:
            key = (entry["shift_date"], entry["project_number"])
            totals_seconds[key] += entry["duration_seconds"]
            totals_names[entry["project_number"]] = entry["project_name"]
            totals_shift[key] = entry["shift"]
            totals_timezone[key] = entry["timezone"]

        wb = openpyxl.load_workbook(self.current_database_path)

        # Heal headers on older databases created before the Shift/Time Zone
        # columns existed.
        self._ensure_headers(wb["Daily Log"], ["Project Name", "Project Number", "Start", "End", "Duration", "Shift", "Time Zone"])
        self._ensure_headers(wb["Project Totals"], ["Project Name", "Project Number", "Total", "Shift Date", "Shift", "Time Zone"])

        ws_log = wb["Daily Log"]
        for entry in self.session_log:
            ws_log.append([
                entry["project_name"], entry["project_number"], entry["start"], entry["end"],
                format_hhmmss(entry["duration_seconds"]), entry["shift"], entry["timezone"]
            ])

        ws_totals = wb["Project Totals"]

        # Look up any row that already exists for (shift date, project number),
        # so a second "Update Database" for the same project on the same shift
        # day adds to it instead of creating a duplicate row.
        existing_rows = {}
        for row_idx, row in enumerate(ws_totals.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[1] is None:
                continue
            _, existing_number, _, existing_shift_date = row[:4]
            existing_rows[(existing_shift_date, existing_number)] = row_idx

        for (shift_date, number), secs in totals_seconds.items():
            key = (shift_date, number)
            if key in existing_rows:
                row_idx = existing_rows[key]
                previous_total = ws_totals.cell(row=row_idx, column=3).value
                combined_seconds = parse_hhmmss(previous_total) + secs
                ws_totals.cell(row=row_idx, column=3, value=format_hhmmss(combined_seconds))
                if not ws_totals.cell(row=row_idx, column=1).value:
                    ws_totals.cell(row=row_idx, column=1, value=totals_names[number])
                if not ws_totals.cell(row=row_idx, column=5).value:
                    ws_totals.cell(row=row_idx, column=5, value=totals_shift[key])
                if not ws_totals.cell(row=row_idx, column=6).value:
                    ws_totals.cell(row=row_idx, column=6, value=totals_timezone[key])
            else:
                ws_totals.append([
                    totals_names[number], number, format_hhmmss(secs), shift_date,
                    totals_shift[key], totals_timezone[key]
                ])

        wb.save(self.current_database_path)

        messagebox.showinfo(
            title="Update Database",
            message=f"Saved {len(self.session_log)} entr{'y' if len(self.session_log) == 1 else 'ies'} to the database."
        )

        # Committed to the archive - clear the in-memory session so it isn't saved twice
        self.session_log = []

    # -- "View Day's Data": read-only look at today's entries, pulled straight from
    #    the in-memory session_log (auto-filled on every stop-click). Update Database
    #    plays no part in this - nothing here touches Excel. --
    def on_view_days_data(self):
        # "Today" here means the CURRENT shift date under the active profile,
        # not the literal calendar date - so a 3rd-shift worker checking this
        # at 2am still sees the shift they're in the middle of, not an empty
        # "today" with last night's entries missing.
        tz = ZoneInfo(self.profile.get("timezone", "America/New_York"))
        shift = self.profile.get("shift", 1)
        now_local = datetime.datetime.now(datetime.timezone.utc).astimezone(tz)
        shift_date = compute_shift_date(now_local, shift).isoformat()

        todays_entries = [e for e in self.session_log if e.get("shift_date") == shift_date]

        win = tk.Toplevel(self)
        win.title(f"Data for Shift Date {shift_date}")
        win.configure(bg=BG)
        win.geometry("560x600")

        style = ttk.Style(win)
        style.theme_use("clam")
        style.configure("Treeview", background=BG_ALT, fieldbackground=BG_ALT, foreground=FG)
        style.configure("Treeview.Heading", background=BG, foreground=FG)

        # -- Section: today's log entries --
        tk.Label(
            win, text="Today's Log", font=("Segoe UI", 11, "bold"), bg=BG, fg=FG
        ).pack(anchor="w", padx=10, pady=(10, 4))

        log_columns = ("project_name", "project_number", "start", "end", "duration")
        log_headers = ("Project Name", "Project Number", "Start", "End", "Duration")

        log_tree = ttk.Treeview(win, columns=log_columns, show="headings", height=8)
        for col, label in zip(log_columns, log_headers):
            log_tree.heading(col, text=label)
            log_tree.column(col, width=100, anchor="center")
        log_tree.pack(fill="x", padx=10)

        if todays_entries:
            for entry in todays_entries:
                log_tree.insert("", "end", values=(
                    entry["project_name"], entry["project_number"], entry["start"], entry["end"],
                    format_hhmmss(entry["duration_seconds"])
                ))
        else:
            tk.Label(
                win, text="No entries recorded for today yet.",
                bg=BG, fg=FG_MUTED
            ).pack(anchor="w", padx=10, pady=4)

        ttk.Separator(win, orient="horizontal").pack(fill="x", padx=10, pady=12)

        # -- Section: project totals --
        tk.Label(
            win, text="Project Totals", font=("Segoe UI", 11, "bold"), bg=BG, fg=FG
        ).pack(anchor="w", padx=10, pady=(0, 4))

        totals_seconds = defaultdict(int)
        totals_names = {}
        for entry in todays_entries:
            key = entry["project_number"]
            totals_seconds[key] += entry["duration_seconds"]
            totals_names[key] = entry["project_name"]

        totals_columns = ("project_name", "project_number", "total", "shift_date")
        totals_headers = ("Project Name", "Project Number", "Total for Shift", "Shift Date")

        totals_tree = ttk.Treeview(win, columns=totals_columns, show="headings", height=6)
        for col, label in zip(totals_columns, totals_headers):
            totals_tree.heading(col, text=label)
            totals_tree.column(col, width=130, anchor="center")
        totals_tree.pack(fill="x", padx=10, pady=(0, 10))

        if totals_seconds:
            for number, secs in sorted(totals_seconds.items(), key=lambda kv: totals_names[kv[0]]):
                totals_tree.insert("", "end", values=(totals_names[number], number, format_hhmmss(secs), shift_date))
        else:
            tk.Label(
                win, text="No totals to show yet.",
                bg=BG, fg=FG_MUTED
            ).pack(anchor="w", padx=10, pady=4)

        # -- Overhead Total: 8 hours minus everything logged today. Goes negative
        #    (shown as -HH:MM:SS) if today's total already exceeds 8 hours. --
        overhead_seconds = (8 * 3600) - sum(totals_seconds.values())
        tk.Label(
            win, text=f"Overhead Total: {format_hhmmss(overhead_seconds)}",
            font=("Segoe UI", 11, "bold"), bg=BG, fg=FG
        ).pack(anchor="w", padx=10, pady=(4, 10))

    # -- Creates a new TimerBox. Input format is "Name, Project Number" - the
    #    project number must be at least 6 digits, since it's the identifier the
    #    JSON sidecar and Excel both key off of. --
    def on_create_project_timer(self):
        user_input = simpledialog.askstring(
            title="New Project Timer",
            prompt="Enter as: Name, Project Number\n"
                   "(project number must be at least 6 digits)\n"
                   "e.g. Kitchen Remodel, 123456"
        )

        if user_input is None:
            return  # Cancelled
        if user_input.strip() == "":
            return  # Blank input, ignore

        if "," not in user_input:
            messagebox.showerror(
                title="Invalid Format",
                message="Please enter using the format: Name, Project Number\n"
                        "e.g. Kitchen Remodel, 123456"
            )
            return

        name_part, _, number_part = user_input.partition(",")
        project_name = name_part.strip()
        project_number = number_part.strip().replace(" ", "")

        if not project_name:
            messagebox.showerror(
                title="Invalid Format",
                message="Please include a project name before the comma."
            )
            return

        if not project_number.isdigit() or len(project_number) < 6:
            messagebox.showerror(
                title="Invalid Project Number",
                message="Project number must be at least 6 digits (numbers only)."
            )
            return

        if any(t.project_number == project_number for t in self.timers):
            messagebox.showerror(
                title="Duplicate Project Number",
                message=f"A timer for project number {project_number} already exists."
            )
            return

        timer = TimerBox(
            self.timers_container, project_name, project_number,
            on_delete=self.remove_timer,
            on_log_entry=self.add_log_entry,
            on_state_change=self.handle_timer_state_change
        )
        self.timers.append(timer)
        self._relayout_timers()

        # Keep the sidecar's name mapping current
        self.project_names[project_number] = project_name
        self._save_sidecar()

    # -- Called by a TimerBox every time a stop-click produces a finished entry.
    #    This is the auto-fill: session_log updates the instant a timer is stopped.
    #    This is also the ONE place Shift/TimeZone/ShiftDate get resolved and baked
    #    into an entry, using whatever the profile currently says - so if the
    #    profile changes later, already-logged entries keep the values they were
    #    created with instead of silently being reinterpreted. --
    def add_log_entry(self, entry):
        shift = self.profile.get("shift", 1)
        tz_name = self.profile.get("timezone", "America/New_York")
        tz = ZoneInfo(tz_name)

        start_local = entry.pop("start_utc").astimezone(tz)
        end_local = entry.pop("end_utc").astimezone(tz)

        entry["start"] = start_local.strftime("%Y-%m-%d %H:%M:%S")
        entry["end"] = end_local.strftime("%Y-%m-%d %H:%M:%S")
        entry["shift"] = shift
        entry["timezone"] = tz_name
        # Shift date is anchored to the LOCAL start time - a shift is attributed
        # to the day it started on, even if it runs past midnight.
        entry["shift_date"] = compute_shift_date(start_local, shift).isoformat()

        self.session_log.append(entry)

    # -- Called by a TimerBox the instant it starts or stops. While any timer is
    #    running, every other button in every section is disabled, and every other
    #    timer is locked - only the running timer's own "Stop" stays clickable. --
    def handle_timer_state_change(self, active_timer, running):
        if running:
            self._lock_everything_except(active_timer)
        else:
            # Only unlock once nothing else is still running
            if not any(t.running for t in self.timers):
                self._unlock_everything()

    def _lock_everything_except(self, active_timer):
        for btn in (
            self.create_db_btn, self.load_db_btn, self.view_days_btn,
            self.update_db_btn, self.options_mb, self.create_timer_btn, self.delete_timer_btn
        ):
            btn.config(state=tk.DISABLED)

        for timer in self.timers:
            if timer is not active_timer:
                timer.locked = True

    def _unlock_everything(self):
        # Respect delete mode's own lock if it's somehow still active
        if not self.delete_mode:
            for btn in (self.create_db_btn, self.load_db_btn, self.view_days_btn, self.update_db_btn, self.options_mb):
                btn.config(state=tk.NORMAL)
            self._refresh_timer_controls()  # Create/Delete Timer respect the "database loaded?" rule too
            for timer in self.timers:
                timer.locked = False

    # -- Toggles delete mode: shows/hides the "x" on every timer box, and disables
    #    every other button so "Done Deleting" is the only way out --
    def on_delete_timer(self):
        self.delete_mode = not self.delete_mode

        for timer in self.timers:
            if self.delete_mode:
                timer.show_delete_x()
            else:
                timer.hide_delete_x()

        # Stored as 'self' attributes so now reachable from 'on_delete_timer()'
        other_state = tk.DISABLED if self.delete_mode else tk.NORMAL
        self.create_db_btn.config(state=other_state)
        self.load_db_btn.config(state=other_state)
        self.view_days_btn.config(state=other_state)
        self.create_timer_btn.config(state=other_state)
        self.update_db_btn.config(state=other_state)
        self.options_mb.config(state=other_state)

        self.delete_timer_btn.config(text="Done Deleting" if self.delete_mode else "Delete Timer")

    # -- Called by a TimerBox once the user confirms deletion --
    def remove_timer(self, timer):
        if timer in self.timers:
            self.timers.remove(timer)
        timer.destroy()
        self._relayout_timers()

    # -- Arranges all current TimerBoxes into a grid inside timers_container --
    def _relayout_timers(self):
        for widget in self.timers_container.winfo_children():
            widget.grid_forget()

        for index, timer in enumerate(self.timers):
            row = index // self.COLUMNS
            col = index % self.COLUMNS
            timer.grid(row=row, column=col, padx=8, pady=8)


def _check_tzdata():
    """Windows doesn't ship an IANA timezone database the way Linux/macOS do,
    so `zoneinfo` needs the 'tzdata' PyPI package to look up zones like
    'America/Chicago'. Fails fast with a clear popup + fix instead of a raw
    traceback if it's missing."""
    try:
        ZoneInfo("UTC")
    except Exception:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(
            title="Missing Time Zone Data",
            message="This program needs the 'tzdata' package to handle timezones "
                    "correctly (Windows doesn't include it by default).\n\n"
                    "Run this in a terminal, then restart the program:\n\n"
                    "    pip install tzdata"
        )
        root.destroy()
        raise SystemExit(1)

if __name__ == "__main__":
    _check_tzdata()
    mainWindow = MainWindow()
    # Begins an event loop, listening for OS-level events and triggering callback functions
    mainWindow.mainloop()